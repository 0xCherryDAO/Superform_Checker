from asyncio import run, set_event_loop_policy, sleep, create_task, gather
import asyncio
import logging
import sys
import random

from rich.console import Console
from loguru import logger
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from config import MOBILE_PROXY, ROTATE_IP, PAUSE_BETWEEN_WALLETS
from src.checker.checker import get_superform_data
from src.utils.data.helper import proxies, addresses
from src.utils.proxy_manager import Proxy

if sys.platform == 'win32':
    set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())

logging.getLogger("asyncio").setLevel(logging.CRITICAL)


async def prepare_proxy(proxy: str) -> Proxy | None:
    if proxy:
        change_link = None
        if MOBILE_PROXY:
            proxy_url, change_link = proxy.split('|')
        else:
            proxy_url = proxy

        proxy = Proxy(proxy_url=f'http://{proxy_url}', change_link=change_link)

        if ROTATE_IP and MOBILE_PROXY:
            await proxy.change_ip()

        return proxy


async def process_address(address: str, proxy_index: int) -> list:
    proxy = await prepare_proxy(proxies[proxy_index])
    address, xp_final, super_fren_data = await get_superform_data(address, proxy)
    row = [
        address,
        xp_final,
        super_fren_data.get("SuperChad", "N/A"),
        super_fren_data.get("SuperApe", "N/A"),
        super_fren_data.get("SuperWhale", "N/A"),
        super_fren_data.get("SuperFrog", "N/A"),
        super_fren_data.get("SuperDino", "N/A"),
        super_fren_data.get("SuperSnake", "N/A"),
        super_fren_data.get("SuperHog", "N/A"),
    ]
    return row


async def main() -> None:
    workbook = Workbook()
    console = Console()
    sheet = workbook.active
    sheet.title = "Superform Data"

    headers = [
        "address",
        "total points",
        "SuperChad",
        "SuperApe",
        "SuperWhale",
        "SuperFrog",
        "SuperDino",
        "SuperSnake",
        "SuperHog",
    ]
    sheet.append(headers)

    tasks = []
    with console.status("[bold green]Checking wallets...") as status:
        for index, address in enumerate(addresses):
            console.log(f'Checking {address}...')
            proxy_index = index % len(proxies)
            tasks.append(process_address(address, proxy_index))
            time_to_sleep = random.uniform(PAUSE_BETWEEN_WALLETS[0], PAUSE_BETWEEN_WALLETS[1])
            if time_to_sleep != 0:
                console.log(f'Sleeping {time_to_sleep:.2f} seconds...')
                await sleep(time_to_sleep)
        results = await gather(*tasks)

    for row in results:
        sheet.append(row)
        console.log(f"Wallet {row[0]} checked")

    for column_index, column_cells in enumerate(sheet.columns, start=1):
        max_length = 0
        for cell in column_cells:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        column_letter = get_column_letter(column_index)
        sheet.column_dimensions[column_letter].width = max_length + 2

    workbook.save("superform_data.xlsx")
    logger.success("Excel file 'superform_data.xlsx' has been saved!")


if __name__ == '__main__':
    run(main())
